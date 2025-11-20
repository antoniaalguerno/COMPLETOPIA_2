# --- Imports de Sistema ---
from django.http import HttpResponse 
import re 
import pandas as pd 
from openpyxl import Workbook 
import xlwt 

# --- Imports Adicionales para DRF ---
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status, serializers, permissions
from rest_framework.parsers import MultiPartParser, FormParser

# --- MongoDB Imports ---
from bson.objectid import ObjectId
from bson.errors import InvalidId
from core.mongo import products_collection

# --- Models (Solo Profile permanece en ORM) ---
from registration.models import Profile 
# Nota: Ya no se importa 'from .models import Product'

# -----------------------------------------------------------------
# ⚙️ CONFIGURACIÓN Y UTILIDADES DE PYMONGO
# -----------------------------------------------------------------

def format_product_id(product):
    """Convierte el _id de MongoDB (ObjectId) a 'id' (string) para la respuesta JSON."""
    if product and '_id' in product:
        product['id'] = str(product.pop('_id'))
    return product

# -----------------------------------------------------------------
# 🔹 1. PERMISO PERSONALIZADO (Sin Cambios)
# -----------------------------------------------------------------
class IsAdminUser(permissions.BasePermission):
    """Permiso personalizado para permitir solo a usuarios del grupo 1 (Admin)."""
    message = 'No tienes permisos de administrador para realizar esta acción.'

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        try:
            # Asumimos que Profile sigue usando el ORM/Django
            profile = Profile.objects.get(user=request.user) 
            return profile.group_id == 1
        except Profile.DoesNotExist:
            return False

# -----------------------------------------------------------------
# 🔹 2. SERIALIZER DEL PRODUCTO (Ajustado a Serializer base)
# -----------------------------------------------------------------
class ProductSerializer(serializers.Serializer):
    """Define la estructura para serializar y validar datos de productos de MongoDB."""
    id = serializers.CharField(read_only=True) # Para el _id de MongoDB
    supply_name = serializers.CharField(max_length=255)
    supply_code = serializers.CharField(max_length=50)
    supply_unit = serializers.CharField(max_length=50)
    supply_initial_stock = serializers.IntegerField()
    supply_input = serializers.IntegerField()
    supply_output = serializers.IntegerField()
    supply_total = serializers.IntegerField(read_only=True)

# -----------------------------------------------------------------
# 🔹 3. FUNCIÓN DE VALIDACIÓN (Sin Cambios)
# -----------------------------------------------------------------
def validar_nombre(nombre):
    """Valida que el nombre solo contenga letras y espacios."""
    if not re.match(r'^[a-zA-Z\s]+$', nombre):
        return False
    return True

# -----------------------------------------------------------------
# 🔹 DASHBOARD: LISTAR PRODUCTOS CON BAJO STOCK (LECTURA - FILTRO)
# -----------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def list_low_stock_products(request):
    """
    Obtiene una lista de productos cuyo stock total es menor a 5.
    (ORM: Product.objects.filter(supply_total__lt=5).order_by('supply_name'))
    """
    
    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    try:
        # Pymongo: {'supply_total': {'$lt': 5}}, ordenado por nombre ascendente (1)
        low_stock_cursor = products_collection.find(
            {'supply_total': {'$lt': 5}}
        ).sort('supply_name', 1)
        
        # Convertir cursor a lista y formatear IDs
        low_stock_products = [format_product_id(p) for p in low_stock_cursor]
        
        serializer = ProductSerializer(low_stock_products, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response({'error': f'Error al consultar MongoDB: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# -----------------------------------------------------------------
# 🔹 CRUD: LISTAR (Y BUSCAR) TODOS LOS PRODUCTOS (LECTURA - BÚSQUEDA)
# -----------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def list_products(request):
    """
    Lista todos los productos. Acepta un parámetro 'q' para búsqueda.
    (ORM: Product.objects.all().filter(supply_name__icontains=search_query))
    """

    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        
    search_query = request.GET.get('q', '')
    
    mongo_query = {}
    if search_query:
        # __icontains se traduce a $regex (insensible a mayúsculas/minúsculas)
        mongo_query['supply_name'] = {'$regex': search_query, '$options': 'i'}
        
    try:
        products_cursor = products_collection.find(mongo_query).sort('supply_name', 1)
        
        # Convertir cursor a lista y formatear IDs
        product_list = [format_product_id(p) for p in products_cursor]
        
        serializer = ProductSerializer(product_list, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response({'error': f'Error al consultar MongoDB: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# -----------------------------------------------------------------
# 🔹 CRUD: OBTENER DETALLE DE UN PRODUCTO (LECTURA - POR ID)
# -----------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def get_product_detail(request, product_id):
    """
    Obtiene los detalles de un producto específico por su ID.
    (ORM: Product.objects.get(pk=product_id))
    """
    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        
    try:
        # Convertir el ID de string a ObjectId
        obj_id = ObjectId(product_id)
    except InvalidId:
        return Response({'error': 'ID de producto inválido'}, status=status.HTTP_400_BAD_REQUEST)
        
    # Pymongo: Buscar por _id
    product_dict = products_collection.find_one({'_id': obj_id})
    
    if not product_dict:
        return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
    
    # Formatear el diccionario para el serializer
    data = format_product_id(product_dict)
    
    serializer = ProductSerializer(data) 
    
    data_response = serializer.data
    # Añadimos el código numérico como en tu lógica original
    data_response['supply_code_numeric'] = re.sub(r'\D', '', product_dict.get('supply_code', ''))
    
    return Response(data_response)

# -----------------------------------------------------------------
# 🔹 CRUD: CREAR UN NUEVO PRODUCTO (ESCRITURA - CREACIÓN)
# -----------------------------------------------------------------
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def create_product(request):
    """
    Crea un nuevo producto.
    (ORM: Product.objects.create(...))
    """
    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        
    data = request.data
    
    # --- Validación (Sin Cambios) ---
    supply_name = data.get('supply_name')
    supply_unit = data.get('supply_unit')
    supply_code_num = data.get('supply_code') 
    supply_initial_stock = data.get('supply_initial_stock')
    supply_input = data.get('supply_input', 0) 
    supply_output = data.get('supply_output', 0)

    if not all([supply_name, supply_unit, supply_code_num, supply_initial_stock is not None]):
        return Response({'error': 'Debes ingresar toda la información (nombre, unidad, código, stock inicial)'}, status=status.HTTP_400_BAD_REQUEST)

    if not validar_nombre(supply_name):
        return Response({'error': 'Debes ingresar un nombre válido'}, status=status.HTTP_400_BAD_REQUEST)

    if supply_unit not in ['kg', 'LATA (330 ml)']:
        return Response({'error': 'La unidad no es válida'}, status=status.HTTP_400_BAD_REQUEST)

    if not str(supply_initial_stock).isdigit():
        return Response({'error': 'El stock inicial debe ser un número entero'}, status=status.HTTP_400_BAD_REQUEST)

    supply_code = f'SKU{supply_code_num}'
    if not re.match(r'^SKU\d{4}$', supply_code):
        return Response({'error': 'El código del producto debe tener el formato SKU seguido de 4 dígitos'}, status=status.HTTP_400_BAD_REQUEST)
    # --- Fin Validación ---

    try:
        # Calcular el total y preparar el documento para MongoDB
        initial = int(supply_initial_stock)
        input_val = int(supply_input)
        output_val = int(supply_output)
        total = initial + input_val - output_val
        
        new_product_data = {
            'supply_name': supply_name,
            'supply_code': supply_code,
            'supply_unit': supply_unit,
            'supply_initial_stock': initial,
            'supply_input': input_val,
            'supply_output': output_val,
            'supply_total': total,
        }
        
        # Pymongo: Insertar el nuevo documento
        result = products_collection.insert_one(new_product_data)
        
        # Obtener el documento recién creado (incluye el _id)
        created_product_dict = products_collection.find_one({'_id': result.inserted_id})
        
        formatted_product = format_product_id(created_product_dict)
        
        serializer = ProductSerializer(formatted_product)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# -----------------------------------------------------------------
# 🔹 CRUD: ACTUALIZAR UN PRODUCTO (ESCRITURA - ACTUALIZACIÓN)
# -----------------------------------------------------------------
@api_view(['PUT'])
@permission_classes([IsAuthenticated, IsAdminUser])
def update_product(request, product_id):
    """
    Actualiza un producto existente.
    (ORM: product.save())
    """
    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        
    try:
        obj_id = ObjectId(product_id)
    except InvalidId:
        return Response({'error': 'ID de producto inválido'}, status=status.HTTP_400_BAD_REQUEST)

    # 1. Obtener el producto actual
    product_dict = products_collection.find_one({'_id': obj_id})
    if not product_dict:
        return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    data = request.data
    
    # --- Validación (Sin Cambios) ---
    supply_name = data.get('supply_name')
    supply_code_num = data.get('supply_code') 
    supply_unit = data.get('supply_unit')
    supply_initial_stock = data.get('supply_initial_stock')
    supply_input = data.get('supply_input')
    supply_output = data.get('supply_output')

    if not all([supply_name, supply_unit, supply_code_num, 
                supply_initial_stock is not None, 
                supply_input is not None, 
                supply_output is not None]):
        return Response({'error': 'Debes ingresar toda la información'}, status=status.HTTP_400_BAD_REQUEST)

    if not validar_nombre(supply_name):
        return Response({'error': 'Debes ingresar un nombre de producto válido'}, status=status.HTTP_400_BAD_REQUEST)

    if supply_unit not in ['kg', 'LATA (330 ml)']:
        return Response({'error': 'La unidad no es válida'}, status=status.HTTP_400_BAD_REQUEST)

    if not all(str(val).isdigit() for val in [supply_initial_stock, supply_input, supply_output]):
        return Response({'error': 'El stock inicial, entrada y salida deben ser números enteros'}, status=status.HTTP_400_BAD_REQUEST)

    new_supply_code = f'SKU{supply_code_num}'
    if not re.match(r'^SKU\d{4}$', new_supply_code):
        return Response({'error': 'El código del producto debe tener el formato SKU seguido de 4 dígitos'}, status=status.HTTP_400_BAD_REQUEST)
    # --- Fin Validación ---

    # --- Lógica de negocio (adaptada para dicts) ---
    current_total = product_dict['supply_total']
    current_initial = product_dict['supply_initial_stock']
    
    update_initial_stock = int(supply_initial_stock)
    input_val = int(supply_input)
    output_val = int(supply_output)
    
    # Comprobar si es la primera edición (usando los valores almacenados en MongoDB)
    if str(current_total) != str(current_initial):
        # Si no es la primera edición, el "nuevo stock inicial" es el total anterior
        update_initial_stock = current_total

    # Calcular supply_total
    new_total = update_initial_stock + input_val - output_val

    update_fields = {
        'supply_name': supply_name,
        'supply_code': new_supply_code,
        'supply_unit': supply_unit,
        'supply_initial_stock': update_initial_stock, 
        'supply_input': input_val,
        'supply_output': output_val,
        'supply_total': new_total, 
    }
    
    # Pymongo: Actualizar en MongoDB usando $set
    products_collection.update_one({'_id': obj_id}, {'$set': update_fields})
    
    # Obtener el documento actualizado para la respuesta
    updated_product = products_collection.find_one({'_id': obj_id})
    
    formatted_product = format_product_id(updated_product)
    serializer = ProductSerializer(formatted_product)
    return Response(serializer.data, status=status.HTTP_200_OK)


# -----------------------------------------------------------------
# 🔹 CRUD: ELIMINAR UN PRODUCTO (ESCRITURA - ELIMINACIÓN)
# -----------------------------------------------------------------
@api_view(['DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def delete_product(request, product_id):
    """
    Elimina un producto por su ID.
    (ORM: product.delete())
    """
    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        
    try:
        obj_id = ObjectId(product_id)
    except InvalidId:
        return Response({'error': 'ID de producto inválido'}, status=status.HTTP_400_BAD_REQUEST)
        
    # Pymongo: Eliminar por _id
    result = products_collection.delete_one({'_id': obj_id})
    
    if result.deleted_count == 0:
        return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
    return Response({'message': 'Producto eliminado correctamente'}, status=status.HTTP_204_NO_CONTENT)


# -----------------------------------------------------------------
# 🔹 EXTRAS: IMPORTACIÓN / EXPORTACIÓN
# -----------------------------------------------------------------

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def download_template(request):
    """Descarga la plantilla Excel para la carga masiva. (Sin cambios de BD)"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="archivo_importacion_producto.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = 'carga_masiva'
    columns = ['Producto', 'Código', 'Unidad', 'Stock inicial']
    ws.append(columns)
    example_data = ['ej: Palta', 'ej: SKU1111', 'ej: kg/LATA (330 ml)', 'ej: 10']
    ws.append(example_data)
    wb.save(response)
    return response

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser]) 
def upload_bulk_products(request):
    """
    Procesa un archivo Excel de carga masiva de productos.
    (ORM: Product.objects.create) -> Pymongo: insert_many
    """
    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    if 'myfile' not in request.FILES:
        return Response({'error': 'No se encontró ningún archivo "myfile"'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        data = pd.read_excel(request.FILES['myfile'], engine='openpyxl', skiprows=1)
        df = pd.DataFrame(data)
        bulk_operations = []
        acc = 0
        
        for item in df.itertuples():
            # Limpieza y validación
            supply_name = str(item[1])
            supply_code = str(item[2])
            supply_unit = str(item[3])
            
            try:
                # Asegurarse de que el stock es convertible a entero
                supply_initial_stock = int(item[4])
            except (ValueError, TypeError):
                raise ValueError(f'El stock inicial "{item[4]}" debe ser un número entero')
            
            if not validar_nombre(supply_name):
                raise ValueError(f'El nombre del insumo "{supply_name}" solo puede contener letras')
            if supply_unit not in ['kg', 'LATA (330 ml)']:
                raise ValueError(f'La unidad "{supply_unit}" no es válida')
            
            # Preparar documento para inserción
            new_product = {
                'supply_name': supply_name,
                'supply_code': supply_code,
                'supply_unit': supply_unit,
                'supply_initial_stock': supply_initial_stock,
                'supply_output': 0,
                'supply_input': 0,
                'supply_total': supply_initial_stock,
            }
            bulk_operations.append(new_product)
            acc += 1

        if bulk_operations:
            # Pymongo: Inserción masiva para mejor rendimiento
            products_collection.insert_many(bulk_operations)
        
        return Response({'message': f'Carga masiva finalizada, se importaron {acc} registros'}, status=status.HTTP_201_CREATED)
    
    except Exception as e:
        # Captura errores de validación de los datos en el Excel o de inserción
        return Response({'error': f'Error al procesar el archivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def download_report_all(request):
    """
    Descarga un reporte en Excel con todos los productos.
    (ORM: Product.objects.all().order_by('supply_name'))
    """
    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        
    try:
        # Pymongo: find() con sort
        productos_cursor = products_collection.find().sort('supply_name', 1)

        style_2 = xlwt.easyxf('font: name Time New Roman, color-index black; font: bold on')
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        response=HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ListaProductos.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Productos')
        row_num = 0
        columns = ['Producto', 'Código', 'Unidad', 'Stock Total'] # Añadido stock total para más utilidad
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num] ,font_style)

        for row in productos_cursor:
            row_num += 1
            ws.write(row_num, 0, row.get('supply_name', ''), style_2)
            ws.write(row_num, 1, row.get('supply_code', ''), style_2) 
            ws.write(row_num, 2, row.get('supply_unit', ''), style_2)
            ws.write(row_num, 3, row.get('supply_total', 0), style_2)
        
        wb.save(response)
        return response
    except Exception as e:
        return Response({'error': f'Se produjo un error al generar el archivo Excel: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def download_report_filtered(request):
    """
    Descarga un reporte en Excel filtrado por nombre de producto.
    (ORM: Product.objects.filter(supply_name__icontains=producto_query))
    """
    if products_collection is None:
        return Response({'error': 'Error de conexión a la base de datos'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        
    try:
        producto_query = request.GET.get('q', '') 
        if not producto_query:
            return Response({'error': 'El filtro "q" (producto) no puede estar vacío'}, status=status.HTTP_400_BAD_REQUEST)

        # Pymongo: Filtro por nombre (case-insensitive regex)
        mongo_query = {'supply_name': {'$regex': producto_query, '$options': 'i'}}
        productos_cursor = products_collection.find(mongo_query)
        
        productos_array = list(productos_cursor)

        if not productos_array:
            return Response({'error': 'No existe producto con la cadena buscada'}, status=status.HTTP_404_NOT_FOUND)

        # (Tu lógica de creación de Excel...)
        style_2 = xlwt.easyxf('font: name Times New Roman, color-index black, bold on')
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="ReporteProductos_{producto_query}.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Productos')
        row_num = 0
        columns = ['Producto', 'Código', 'Unidad', 'Stock Total']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], style_2)

        for row in productos_array:
            row_num += 1
            ws.write(row_num, 0, row.get('supply_name', ''), style_2)
            ws.write(row_num, 1, row.get('supply_code', ''), style_2)
            ws.write(row_num, 2, row.get('supply_unit', ''), style_2)
            ws.write(row_num, 3, row.get('supply_total', 0), style_2)
        
        wb.save(response)
        return response
    except Exception as e:
        return Response({'error': f'Error al generar el reporte: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)